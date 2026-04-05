import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from agent_core import PerformanceMatrixAgent
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

st.set_page_config(
    page_title="AI-агент эффективности сотрудников",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
    * { font-family: 'Inter', sans-serif; }
    .main-header {
        background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
        padding: 2.5rem;
        border-radius: 1.5rem;
        margin-bottom: 2rem;
        color: white;
        text-align: center;
    }
    .main-header h1 {
        font-size: 2.5rem;
        font-weight: 800;
        background: linear-gradient(135deg, #fff 0%, #a8c0ff 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .fact-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.25rem;
        border-radius: 1rem;
        color: white;
        text-align: center;
        transition: transform 0.3s ease;
    }
    .fact-card:hover { transform: translateY(-5px); }
    .fact-card h2 { font-size: 2.2rem; font-weight: 800; margin: 0.5rem 0; }
    .prediction-card {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        padding: 1.25rem;
        border-radius: 1rem;
        color: white;
        text-align: center;
        transition: transform 0.3s ease;
    }
    .prediction-card:hover { transform: translateY(-5px); }
    .prediction-card h2 { font-size: 2.2rem; font-weight: 800; margin: 0.5rem 0; }
    .critical-box {
        background: linear-gradient(135deg, #ff7675 0%, #d63031 100%);
        padding: 1.25rem;
        border-radius: 1rem;
        margin: 1rem 0;
        color: white;
    }
    .warning-box {
        background: linear-gradient(135deg, #ffeaa7 0%, #fdcb6e 100%);
        padding: 1.25rem;
        border-radius: 1rem;
        margin: 1rem 0;
    }
    .info-box {
        background: linear-gradient(135deg, #e0eafc 0%, #cfdef3 100%);
        padding: 1.25rem;
        border-radius: 1rem;
        margin: 1rem 0;
    }
    .stButton>button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 0.6rem 2rem;
        border-radius: 0.75rem;
        font-weight: 600;
    }
    .section-title {
        font-size: 1.5rem;
        font-weight: 700;
        margin: 1.5rem 0 1rem 0;
        padding-left: 1rem;
        border-left: 4px solid #667eea;
    }
    .employee-card {
        background: white;
        border-radius: 1rem;
        padding: 1.5rem;
        margin-bottom: 2rem;
        box-shadow: 0 5px 15px -5px rgba(0,0,0,0.1);
    }
    .employee-card h3 { color: #667eea; margin-bottom: 1rem; font-weight: 600; }
    .footer {
        text-align: center;
        padding: 2rem;
        margin-top: 2rem;
        color: #666;
        font-size: 0.85rem;
        border-top: 1px solid #eee;
    }
</style>
""", unsafe_allow_html=True)

agent = PerformanceMatrixAgent()

# Инициализация сессии
if 'combined_df' not in st.session_state:
    st.session_state.combined_df = None
if 'current_department' not in st.session_state:
    st.session_state.current_department = None
if 'selected_dispatchers' not in st.session_state:
    st.session_state.selected_dispatchers = []
if 'dispatcher_confirmed' not in st.session_state:
    st.session_state.dispatcher_confirmed = False
if 'employees_list' not in st.session_state:
    st.session_state.employees_list = []

st.markdown("""
<div class="main-header">
    <h1>🎯 AI-агент эффективности сотрудников</h1>
    <p>Анализ метрик | Прогнозирование | Умные уведомления</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 📊 Управление данными")
    department_type = st.selectbox("Выберите отдел", ["Диспетчерская", "Отдел продаж (в разработке)"])
    
    if department_type == "Отдел продаж (в разработке)":
        st.info("🚧 **В разработке**\n\nБлок анализа отдела продаж будет доступен в следующей версии.")
        st.stop()
    
    input_method = st.radio("Способ ввода", [
        "➕ Один сотрудник",
        "👥 Несколько сотрудников",
        "📁 Загрузить файл(ы) из CRM"
    ])
    
    df_input = None
    
    if input_method == "➕ Один сотрудник":
        with st.form("single_form"):
            name = st.text_input("Имя сотрудника")
            total_leads = st.number_input("Получено лидов", min_value=0, value=0)
            transferred = st.number_input("Передано продажам", min_value=0, value=0)
            failed_calls = st.number_input("Недозвоны", min_value=0, value=0)
            rejected = st.number_input("Отказы", min_value=0, value=0)
            submitted = st.form_submit_button("🚀 Рассчитать", use_container_width=True)
            if submitted and name:
                df_input = pd.DataFrame({
                    'имя': [name],
                    'получено_лидов': [total_leads],
                    'передано_продажам': [transferred],
                    'недозвоны': [failed_calls],
                    'забраковано': [rejected]
                })
                st.session_state.combined_df = df_input
                st.session_state.current_department = "dispatcher"
                st.session_state.dispatcher_confirmed = True
                st.rerun()
    
    elif input_method == "👥 Несколько сотрудников":
        with st.form("multi_form"):
            name = st.text_input("Имя сотрудника")
            total_leads = st.number_input("Получено лидов", min_value=0, value=0)
            transferred = st.number_input("Передано продажам", min_value=0, value=0)
            failed_calls = st.number_input("Недозвоны", min_value=0, value=0)
            rejected = st.number_input("Отказы", min_value=0, value=0)
            add_button = st.form_submit_button("➕ Добавить", use_container_width=True)
            if add_button and name:
                st.session_state.employees_list.append({
                    'имя': name,
                    'получено_лидов': total_leads,
                    'передано_продажам': transferred,
                    'недозвоны': failed_calls,
                    'забраковано': rejected
                })
                st.success(f"✅ Добавлен: {name}")
                st.rerun()
        
        if st.session_state.employees_list:
            st.dataframe(pd.DataFrame(st.session_state.employees_list), use_container_width=True)
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🚀 Рассчитать всех", use_container_width=True):
                    df_input = pd.DataFrame(st.session_state.employees_list)
                    st.session_state.combined_df = df_input
                    st.session_state.current_department = "dispatcher"
                    st.session_state.dispatcher_confirmed = True
                    st.session_state.employees_list = []
                    st.rerun()
            with col2:
                if st.button("🗑️ Очистить список", use_container_width=True):
                    st.session_state.employees_list = []
                    st.rerun()
    
    else:
        st.markdown("### 📁 Загрузка из CRM")
        uploaded_files = st.file_uploader("Выберите файлы", type=["xlsx", "csv"], accept_multiple_files=True)
        
        if uploaded_files:
            if st.button("🔄 Обработать", use_container_width=True):
                all_dfs = []
                for file in uploaded_files:
                    try:
                        if file.name.endswith('xlsx'):
                            df = pd.read_excel(file)
                        else:
                            df = pd.read_csv(file, encoding='utf-8')
                        all_dfs.append(df)
                    except Exception as e:
                        st.error(f"Ошибка: {file.name} - {e}")
                
                if all_dfs:
                    df_input = pd.concat(all_dfs, ignore_index=True)
                    st.session_state.combined_df = df_input
                    st.session_state.current_department = "dispatcher"
                    st.session_state.dispatcher_confirmed = False
                    st.success(f"✅ Загружено {len(uploaded_files)} файлов, {len(df_input)} строк")

# Основная логика анализа
if st.session_state.combined_df is not None and st.session_state.current_department == "dispatcher":
    df = st.session_state.combined_df
    
    if not st.session_state.dispatcher_confirmed:
        columns_map = agent.detect_columns(df)
        if columns_map['маркетолог']:
            all_dispatchers = df[columns_map['маркетолог']].dropna().unique()
            all_dispatchers = [str(d).strip() for d in all_dispatchers if pd.notna(d) and str(d).strip()]
            if all_dispatchers:
                st.markdown('<div class="info-box"><h4>🔍 Подтверждение диспетчеров</h4></div>', unsafe_allow_html=True)
                selected = []
                cols = st.columns(3)
                for i, d in enumerate(all_dispatchers):
                    with cols[i % 3]:
                        if st.checkbox(d, value=True, key=f"disp_{i}"):
                            selected.append(d)
                if st.button("✅ Подтвердить", use_container_width=True):
                    st.session_state.selected_dispatchers = selected
                    st.session_state.dispatcher_confirmed = True
                    st.rerun()
            else:
                st.session_state.dispatcher_confirmed = True
                st.rerun()
        else:
            st.session_state.dispatcher_confirmed = True
            st.rerun()
    
    else:
        with st.spinner("🔍 Анализируем данные..."):
            summary_df, daily_df, landings_df, extra_info = agent.analyze_data(
                df, st.session_state.selected_dispatchers if st.session_state.selected_dispatchers else None
            )
        
        if summary_df.empty:
            st.warning("⚠️ Не удалось проанализировать данные. Проверьте структуру файла.")
        else:
            # Нераспределенные лиды
            unassigned = extra_info.get('unassigned', {})
            unassigned_df = extra_info.get('unassigned_df', pd.DataFrame())
            
            if unassigned and unassigned.get('count', 0) > 0:
                st.markdown(f"""
                <div class="warning-box">
                    <h4>⚠️ НЕ ЗАКРЕПЛЕННЫЕ ЗА ДИСПЕТЧЕРОМ ЛИДЫ</h4>
                    <p><b>{unassigned['count']}</b> лидов ({unassigned['rate']}% от всех) не закреплены за диспетчерами!</p>
                    <p>Эти лиды не были обработаны. Рекомендуется настроить автоматическое распределение.</p>
                </div>
                """, unsafe_allow_html=True)
                
                if not unassigned_df.empty:
                    with st.expander(f"📋 Список незакрепленных лидов ({len(unassigned_df)} шт)", expanded=False):
                        st.dataframe(unassigned_df, use_container_width=True)
            
            issues_df = extra_info.get('issues', pd.DataFrame())
            if not issues_df.empty:
                with st.expander(f"⚠️ У {len(issues_df)} лидов не сменили ответственного при передаче", expanded=False):
                    st.markdown("""
                    <div class="warning-box" style="margin-bottom: 1rem;">
                        <b>📌 Что это значит?</b><br>
                        Статус лида "Передан менеджеру", но поле "Ответственный" осталось с ФИО диспетчера.
                        Это может влиять на статистику и контроль продаж.
                    </div>
                    """, unsafe_allow_html=True)
                    st.dataframe(issues_df, use_container_width=True)
            
            # Сводка по отделу
            st.markdown('<div class="section-title">📊 Сводка по отделу</div>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                actual = summary_df['% переданных'].mean()
                total = summary_df['Передано продажам'].sum()
                st.markdown(f'<div class="fact-card"><h4>📊 ФАКТ</h4><h2>{actual:.1f}%</h2><p>средний % переданных</p><hr><p>📦 Передано: <b>{total:.0f}</b></p></div>', unsafe_allow_html=True)
            with col2:
                pred = summary_df['📈 ПРОГНОЗ: % переданных'].mean()
                pred_total = summary_df['📈 ПРОГНОЗ: передач (шт)'].sum()
                st.markdown(f'<div class="prediction-card"><h4>📈 ПРОГНОЗ (ML)</h4><h2>{pred:.1f}%</h2><p>средний % переданных</p><hr><p>📦 Прогноз: <b>{pred_total:.0f}</b></p></div>', unsafe_allow_html=True)
            with col3:
                failed = summary_df['% недозвонов'].mean()
                st.markdown(f'<div class="fact-card"><h4>📞 НЕДОЗВОНЫ</h4><h2>{failed:.1f}%</h2><p>средний % недозвонов</p><hr><p>🎯 Норма: до 15%</p></div>', unsafe_allow_html=True)
            with col4:
                rejected = summary_df['% отказов'].mean()
                st.markdown(f'<div class="fact-card"><h4>❌ ОТКАЗЫ</h4><h2>{rejected:.1f}%</h2><p>средний % отказов</p><hr><p>🎯 Норма: до 10%</p></div>', unsafe_allow_html=True)
            
            # Анализ лэндов
            if not landings_df.empty:
                st.markdown('<div class="section-title">🌐 Анализ источников (Лэндов)</div>', unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                with col1:
                    fig_land = px.bar(landings_df, x='Лэнд', y='Конверсия %', title="Конверсия по источникам",
                                     text='Конверсия %', color='Оценка', color_discrete_map={'Отлично': '#27ae60', 'Средне': '#f39c12', 'Плохо': '#e74c3c'})
                    fig_land.update_layout(height=400, plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig_land, use_container_width=True, key="landing_chart")
                with col2:
                    st.dataframe(landings_df, use_container_width=True)
            
            # Анализ по сотрудникам
            st.markdown('<div class="section-title">👥 Анализ по сотрудникам</div>', unsafe_allow_html=True)
            display_cols = [c for c in summary_df.columns if c not in ['is_problem']]
            st.dataframe(summary_df[display_cols], use_container_width=True)
            
            # Графики по сотрудникам
            st.markdown('<div class="section-title">📈 Графики по сотрудникам</div>', unsafe_allow_html=True)
            
            for idx, emp in summary_df.iterrows():
                name = emp['Сотрудник']
                st.markdown(f'<div class="employee-card"><h3>👤 {name}</h3>', unsafe_allow_html=True)
                
                col1, col2 = st.columns(2)
                with col1:
                    fig = go.Figure()
                    fig.add_trace(go.Bar(x=['Текущий', 'Прогноз'], y=[emp['% переданных'], emp['📈 ПРОГНОЗ: % переданных']],
                                         marker_color=['#667eea', '#f5576c'], text=[f"{emp['% переданных']:.1f}%", f"{emp['📈 ПРОГНОЗ: % переданных']:.1f}%"], textposition='outside'))
                    fig.add_hline(y=70, line_dash="dash", line_color="#27ae60", annotation_text="Норма 70%")
                    fig.update_layout(title="% переданных: факт vs прогноз", height=350, plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig, use_container_width=True, key=f"bar_chart_{idx}")
                
                with col2:
                    metrics = []
                    values = []
                    if emp['% недозвонов'] > 15:
                        metrics.append('Недозвоны'); values.append(emp['% недозвонов'])
                    if emp['% отказов'] > 10:
                        metrics.append('Отказы'); values.append(emp['% отказов'])
                    if emp.get('% дублей/повторов', 0) > 5:
                        metrics.append('Дубли/Повторы'); values.append(emp['% дублей/повторов'])
                    
                    if metrics:
                        fig2 = go.Figure()
                        fig2.add_trace(go.Bar(x=metrics, y=values, marker_color='#e74c3c', text=values, textposition='outside'))
                        fig2.update_layout(title="⚠️ Показатели выше нормы", height=350, plot_bgcolor='rgba(0,0,0,0)')
                        st.plotly_chart(fig2, use_container_width=True, key=f"warning_chart_{idx}")
                    else:
                        st.markdown('<div style="background:#27ae60; color:white; padding:1rem; border-radius:1rem; text-align:center; height:350px; display:flex; align-items:center; justify-content:center;"><div><h2>✅</h2><p>Все показатели в норме!</p></div></div>', unsafe_allow_html=True)
                
                if not daily_df.empty and name in daily_df['Сотрудник'].values:
                    emp_daily = daily_df[daily_df['Сотрудник'] == name].sort_values('Дата')
                    fig3 = go.Figure()
                    fig3.add_trace(go.Scatter(x=emp_daily['Дата'], y=emp_daily['% переданных'], mode='lines+markers',
                                             line=dict(color='#667eea', width=3), marker=dict(size=8, color='#f5576c'),
                                             fill='tozeroy', fillcolor='rgba(102,126,234,0.2)'))
                    fig3.add_hline(y=70, line_dash="dash", line_color="#27ae60", annotation_text="Норма 70%")
                    fig3.update_layout(title="Динамика % переданных по дням", xaxis_title="Дата", yaxis_title="% переданных", height=350, hovermode='x unified', plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig3, use_container_width=True, key=f"line_chart_{idx}")
                
                st.markdown('</div>', unsafe_allow_html=True)
            
            # Загруженность
            if 'Лидов в день' in summary_df.columns:
                st.markdown('<div class="section-title">⚡ Загруженность сотрудников</div>', unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                with col1:
                    fig_load = px.bar(summary_df, x='Сотрудник', y='Лидов в день', title="Лидов в день по сотрудникам",
                                     color='Загруженность', color_discrete_map={'Норма ✅': '#27ae60', 'Недогруз ⚠️': '#f39c12', 'Критическая недогрузка 🔴': '#e74c3c'}, text='Лидов в день')
                    fig_load.add_hline(y=20, line_dash="dash", line_color="#2196f3", annotation_text="Норма 20+")
                    fig_load.update_layout(height=400, plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig_load, use_container_width=True, key="load_chart")
                with col2:
                    load_counts = summary_df['Загруженность'].value_counts()
                    fig_pie = px.pie(values=load_counts.values, names=load_counts.index, title="Распределение загруженности",
                                    color_discrete_sequence=['#27ae60', '#f39c12', '#e74c3c'], hole=0.4)
                    fig_pie.update_layout(height=400)
                    st.plotly_chart(fig_pie, use_container_width=True, key="pie_chart")
            
            # Экспорт в Excel
            st.markdown("---")
            
            def export_to_excel_beautiful(summary_df, landings_df, issues_df, unassigned_df):
                output = io.BytesIO()
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                
                header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                title_font = Font(bold=True, size=14, color="667eea")
                center_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                # Лист 1: Сводка
                ws1 = wb.create_sheet("📊 СВОДКА", 0)
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws1.column_dimensions[col].width = 20
                
                ws1.merge_cells('A1:E1')
                ws1['A1'] = "ОТЧЕТ ПО ЭФФЕКТИВНОСТИ ДИСПЕТЧЕРОВ"
                ws1['A1'].font = Font(bold=True, size=16, color="667eea")
                ws1['A1'].alignment = center_align
                ws1['A3'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                ws1['A3'].font = Font(italic=True, size=10)
                
                row = 5
                if not unassigned_df.empty:
                    ws1.merge_cells(f'A{row}:E{row}')
                    cell = ws1[f'A{row}']
                    cell.value = f"⚠️ НЕ ЗАКРЕПЛЕННЫЕ ЗА ДИСПЕТЧЕРОМ ЛИДЫ: {len(unassigned_df)} шт"
                    cell.fill = yellow_fill
                    cell.font = Font(bold=True)
                    cell.alignment = center_align
                    row += 2
                
                ws1.merge_cells(f'A{row}:E{row}')
                ws1[f'A{row}'] = "ОБЩИЕ ПОКАЗАТЕЛИ ОТДЕЛА"
                ws1[f'A{row}'].font = title_font
                row += 1
                
                headers = ['Показатель', 'Факт', 'Прогноз', 'Норма', 'Статус']
                for c, h in enumerate(headers, 1):
                    cell = ws1.cell(row=row, column=c, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                row += 1
                
                data = [
                    ['% переданных', f"{summary_df['% переданных'].mean():.1f}%", 
                     f"{summary_df['📈 ПРОГНОЗ: % переданных'].mean():.1f}%", '70%',
                     '✅ Норма' if summary_df['% переданных'].mean() >= 70 else '⚠️ Проблема'],
                    ['% недозвонов', f"{summary_df['% недозвонов'].mean():.1f}%", '-', '≤15%',
                     '✅ Норма' if summary_df['% недозвонов'].mean() <= 15 else '⚠️ Проблема'],
                    ['% отказов', f"{summary_df['% отказов'].mean():.1f}%", '-', '≤10%',
                     '✅ Норма' if summary_df['% отказов'].mean() <= 10 else '⚠️ Проблемa']
                ]
                
                for d in data:
                    for c, v in enumerate(d, 1):
                        cell = ws1.cell(row=row, column=c, value=v)
                        cell.alignment = center_align
                        cell.border = thin_border
                        if c == 5:
                            if 'Проблема' in v:
                                cell.fill = red_fill
                            else:
                                cell.fill = green_fill
                    row += 1
                
                row += 1
                problem_emps = summary_df[summary_df['is_problem'] == True]
                if len(problem_emps) > 0:
                    ws1.merge_cells(f'A{row}:E{row}')
                    ws1[f'A{row}'] = "⚠️ СОТРУДНИКИ С ПРОБЛЕМАМИ"
                    ws1[f'A{row}'].font = Font(bold=True, color="e74c3c")
                    row += 1
                    
                    for _, emp in problem_emps.iterrows():
                        issues_list = []
                        if emp['% переданных'] < 70:
                            issues_list.append(f"низкая конверсия {emp['% переданных']}%")
                        if emp['% недозвонов'] > 15:
                            issues_list.append(f"недозвоны {emp['% недозвонов']}%")
                        if emp['% отказов'] > 10:
                            issues_list.append(f"отказы {emp['% отказов']}%")
                        
                        ws1[f'A{row}'] = f"• {emp['Сотрудник']}: {', '.join(issues_list)}"
                        ws1[f'A{row}'].fill = yellow_fill
                        row += 1
                
                # Лист 2: Сотрудники
                ws2 = wb.create_sheet("👥 СОТРУДНИКИ")
                cols_to_show = [c for c in summary_df.columns if c not in ['is_problem']]
                for c, h in enumerate(cols_to_show, 1):
                    cell = ws2.cell(row=1, column=c, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                for r, row_data in enumerate(summary_df[cols_to_show].values, 2):
                    for c, val in enumerate(row_data, 1):
                        ws2.cell(row=r, column=c, value=val)
                
                for c in range(1, len(cols_to_show) + 1):
                    max_len = 0
                    for r in range(1, len(summary_df) + 2):
                        val = ws2.cell(row=r, column=c).value
                        if val:
                            max_len = max(max_len, len(str(val)))
                    ws2.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 20)
                
                if len(summary_df) > 0:
                    chart = BarChart()
                    chart.title = "% переданных по сотрудникам"
                    chart.y_axis.title = "Процент"
                    chart.x_axis.title = "Сотрудник"
                    chart.width = 18
                    chart.height = 12
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showVal = True
                    
                    data_col = cols_to_show.index('% переданных') + 1
                    data_ref = Reference(ws2, min_col=data_col, min_row=1, max_row=len(summary_df) + 1)
                    cat_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(summary_df) + 1)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cat_ref)
                    
                    start_row = len(summary_df) + 4
                    ws2.add_chart(chart, f"A{start_row}")
                
                # Лист 3: Лэнды
                if not landings_df.empty:
                    ws3 = wb.create_sheet("🌐 АНАЛИЗ ЛЭНДОВ")
                    for c, h in enumerate(landings_df.columns, 1):
                        cell = ws3.cell(row=1, column=c, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    
                    for r, row_data in enumerate(landings_df.values, 2):
                        for c, val in enumerate(row_data, 1):
                            ws3.cell(row=r, column=c, value=val)
                    
                    for c in range(1, len(landings_df.columns) + 1):
                        max_len = 0
                        for r in range(1, len(landings_df) + 2):
                            val = ws3.cell(row=r, column=c).value
                            if val:
                                max_len = max(max_len, len(str(val)))
                        ws3.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 18)
                    
                    if len(landings_df) > 0:
                        chart_land = BarChart()
                        chart_land.title = "Конверсия по лэндам"
                        chart_land.y_axis.title = "Процент"
                        chart_land.x_axis.title = "Лэнд"
                        chart_land.width = 18
                        chart_land.height = 12
                        chart_land.dataLabels = DataLabelList()
                        chart_land.dataLabels.showVal = True
                        
                        data_ref = Reference(ws3, min_col=landings_df.columns.get_loc('Конверсия %') + 1,
                                           min_row=1, max_row=len(landings_df) + 1)
                        cat_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(landings_df) + 1)
                        chart_land.add_data(data_ref, titles_from_data=True)
                        chart_land.set_categories(cat_ref)
                        
                        start_row = len(landings_df) + 4
                        ws3.add_chart(chart_land, f"A{start_row}")
                
                # Лист 4: Незакрепленные лиды
                if not unassigned_df.empty:
                    ws4 = wb.create_sheet("⚠️ НЕ ЗАКРЕПЛЕННЫЕ ЛИДЫ")
                    for c, h in enumerate(unassigned_df.columns, 1):
                        cell = ws4.cell(row=1, column=c, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    
                    for r, row_data in enumerate(unassigned_df.values, 2):
                        for c, val in enumerate(row_data, 1):
                            ws4.cell(row=r, column=c, value=val)
                    
                    for c in range(1, len(unassigned_df.columns) + 1):
                        max_len = 0
                        for r in range(1, len(unassigned_df) + 2):
                            val = ws4.cell(row=r, column=c).value
                            if val:
                                max_len = max(max_len, len(str(val)))
                        ws4.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 25)
                
                # Лист 5: Проблемные лиды
                if not issues_df.empty:
                    ws5 = wb.create_sheet("⚠️ ПРОБЛЕМНЫЕ ЛИДЫ")
                    for c, h in enumerate(issues_df.columns, 1):
                        cell = ws5.cell(row=1, column=c, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    
                    for r, row_data in enumerate(issues_df.values, 2):
                        for c, val in enumerate(row_data, 1):
                            ws5.cell(row=r, column=c, value=val)
                    
                    for c in range(1, len(issues_df.columns) + 1):
                        max_len = 0
                        for r in range(1, len(issues_df) + 2):
                            val = ws5.cell(row=r, column=c).value
                            if val:
                                max_len = max(max_len, len(str(val)))
                        ws5.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 25)
                
                wb.save(output)
                return output
            
            excel_file = export_to_excel_beautiful(summary_df, landings_df, issues_df, unassigned_df)
            st.download_button(
                label="📥 Скачать Excel отчет (с графиками)",
                data=excel_file,
                file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.markdown("---")
            if st.button("🔄 Начать заново", use_container_width=True):
                st.session_state.combined_df = None
                st.session_state.dispatcher_confirmed = False
                st.session_state.selected_dispatchers = []
                st.session_state.employees_list = []
                st.rerun()

st.markdown(f"""
<div class="footer">
    <p>🤖 AI-агент эффективности сотрудников | Версия 1.0 | Прогноз на основе линейной регрессии</p>
    <p>© 2024 | Дата отчета: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
</div>
""", unsafe_allow_html=True)